from PyQt5 import uic, QtWidgets
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QPushButton, QMessageBox
from PyQt5.uic import loadUi
from PyQt5.QtCore import pyqtSignal, QThread, QObject
from PyQt5.uic.uiparser import QtWidgets
import sys
from DataHandler import DataHandler
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font
from LaptopUitleenUI import Ui_LaptopUitleenSysteem

class Ui(QMainWindow):
    def __init__(self):
        '''
        Start de UI        
        '''
        super(Ui, self).__init__()
        self.ui = Ui_LaptopUitleenSysteem()
        self.ui.setupUi(self)
        # uic.loadUi(Ui_LaptopUitleenSysteem, self)
        self.show()
        self.ui.RegistreerUitleenButton.clicked.connect(self.RegistreerUitleenButtonClicked)
        self.ui.ZoekUitleenButton.clicked.connect(self.ZoekUitleenButtonClicked)
        self.ui.ExporteerNaarExcelButton.clicked.connect(self.ExporteerNaarExcelButtonClicked)
        self.ui.ToggleFullscreenButton.clicked.connect(self.ToggleFullscreenButtonClicked)
        self.datahandler = DataHandler()

    def WarnMSG(self, Errortype):
        '''
        Geeft een waarschuwing weer met uitleg als er fouten worden opgevangen

        Args:
            type: Het type waarschuwing        
        '''
        msg = QMessageBox()
        msg.setWindowTitle("Er is iets fout gegaan")
        messagewarnings = {"EmptyPersoonBarcode" : "Het veld Leerling/Leraar barcode is leeg!",
                           "EmptyLaptopBarcode" : "Het veld Laptop Barcode is leeg!",
                           "EmptySearchBarcode" : "Het veld Zoek op Barcode is leeg!",
                           "BothBarcodeEmpty" : "De velden Leerling/Leraar barcode en Laptop Barcode zijn leeg!",
                           "NoResult" : "Er is geen resultaat met de barcode gevonden",
                           "SaveError" : "Het bestand opslaan is niet gelukt"}
        msg.setText(messagewarnings[Errortype])
        msg.setIcon(QMessageBox.Warning)
        msg.exec_()
        


    def ResultMSG(self, Resultaat):
        '''
        Geeft een window weer met het zoekresultaat

        Args:
            Resultaat: Het weer te geven resultaat
        '''
        msg = QMessageBox()
        msg.setWindowTitle("Resultaat")
        msg.setText(Resultaat)
        msg.setIcon(QMessageBox.Information)
        msg.exec_()


    def RegistreerUitleenButtonClicked(self):
        '''
        Registreet de uitleen van een laptop.
        Neemt de input van de textboxes Leerling/Leraar Barcode en Laptop Barcode en verstuurt deze naar de data handler die het verder afhandelt.
        '''
        PersoonBarcode = self.GetPersonBarcode()
        LaptopBarcode = self.GetLaptopBarcode()
        Datum = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        if PersoonBarcode == "" and LaptopBarcode == "":
            self.WarnMSG("BothBarcodeEmpty")
            return
        elif PersoonBarcode == "":
            self.WarnMSG("EmptyPersoonBarcode")
            return
        elif LaptopBarcode == "":
            self.WarnMSG("EmptyLaptopBarcode")
            return
        self.datahandler.DataWrite(PersoonBarcode, LaptopBarcode, Datum)
        self.ClearUitleenBoxes()
        

    def ZoekUitleenButtonClicked(self):
        '''
        Zoekt een uitleen op basis van een ingevoerde barcode. Dit kan een Gebruiker of een Laptop barcode zijn.
        '''
        ZoekBarcode = self.GetSearchBarcode()
        if ZoekBarcode == "":
            self.WarnMSG("EmptySearchBarcode")
            return
        
        Resultaat = self.datahandler.DataSearch(ZoekBarcode)
        if Resultaat == []:
            self.WarnMSG("NoResult")
            return
        
        ResultaatString = f"ID: {Resultaat[0]}, Leerling/Leraar: {Resultaat[1]}, Laptop: {Resultaat[2]}, Uitgifte Datum: {Resultaat[3]}"
        self.ResultMSG(ResultaatString)

        self.ClearZoekBoxes()
        

    def ExporteerNaarExcelButtonClicked(self):
        '''
        Exporteert de gegevens naar een excel bestand wat makkelijk leesbaar is
        '''
        Registraties = self.datahandler.DataDump()
        self.MaakExcel(Registraties)
        

    def ToggleFullscreenButtonClicked(self):
        '''
        Switched het scherm naar fullscreen (en terug)
        '''
        if self.isFullScreen() == True:
            self.showNormal()
            self.ui.ToggleFullscreenButton.setText("Toggle Fullscreen")
        else:
            self.showFullScreen()
            self.ui.ToggleFullscreenButton.setText("Toggle Windowed")


    def GetPersonBarcode(self):
        '''
        Functie om Leerling/Leraar barcode uit de textbox te halen

        Return:
            De text uit PersoonBarcodeLineEdit
        '''
        return self.ui.PersoonBarcodeLineEdit.text()


    def GetLaptopBarcode(self):
        '''
        Functie om Laptop barcode uit de textbox te halen

        Return:
            De text uit LaptopBarcodeLineEdit
        '''
        return self.ui.LaptopBarcodeLineEdit.text()


    def GetSearchBarcode(self):
        '''
        Functie om te zoeken barcode uit de textbox te halen

        Return
            De text uit ZoekLaptopBarcodeLineEdit
        '''
        return self.ui.ZoekLaptopBarcodeLineEdit.text()


    def ClearUitleenBoxes(self):
        '''
        Functie om invoerboxes leeg te maken nadat invoer is gebruikt of opgeslagen
        '''
        self.ui.PersoonBarcodeLineEdit.setText("")
        self.ui.LaptopBarcodeLineEdit.setText("")


    def ClearZoekBoxes(self):
        '''
        Functie om invoerbox leeg te maken nadat informatie is opgezocht
        '''
        self.ui.ZoekLaptopBarcodeLineEdit.setText("")


    def MaakExcel(self, Registraties):
        '''
        Maakt het Excel bestand van de uitgeleende laptops, en vult deze met de data uit de database

        Args:
            Registraties: Alle informatie uit de database
        '''
        DatumVandaag = datetime.now().strftime('%Y-%m-%d')
        Bestandsnaam = f"Uitleen_{DatumVandaag}.xlsx"

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Uitleen"

        header_fill = PatternFill(start_color="00C0C0C0", end_color="00C0C0C0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        headers = ["ID", "Leerling/Leraar Barcode", "Laptop Barcode",   "Uitgifte Datum", "Inleverdatum"]
        sheet.append(headers)
        for col in range(1, len(headers) + 1):
            sheet.cell(row=1, column=col).fill = header_fill
            sheet.cell(row=1, column=col).font = header_font

        for index, registratie in enumerate(Registraties, start=2):
            sheet.append(registratie[:-1])

        if registratie[5] == 1:
            sheet.row_dimensions[index].fill = PatternFill(start_color="00FF0000", end_color="00FF0000", fill_type="solid")
        elif registratie[4]:
            sheet.row_dimensions[index].fill = PatternFill(start_color="00CCFFCC", end_color="00CCFFCC", fill_type="solid")

        sheet.column_dimensions['B'].width = 30
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
    	
        try:
            workbook.save(Bestandsnaam)
            self.ResultMSG(f"Gegevens succesvol geëxporteerd naar {Bestandsnaam}")
        except Exception as e:
            self.WarnMSG("SaveError")
            

        

if __name__ == "__main__":
    app = QApplication(sys.argv)  # Create an instance of QtWidgets.QApplication
    window = Ui()  # Create an instance of our class
    app.exec_()  # Start the application    